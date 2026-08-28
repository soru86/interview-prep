from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from recruiter_agent.models import TrackerRow
from recruiter_agent.utils.logging import get_logger

log = get_logger(__name__)

DEFAULT_COLUMNS = [
    "Recruiter Name",
    "Contact Email",
    "Contact Phone",
    "Company",
    "Role Applied For",
    "Match Score",
    "Date of First Reply",
    "Email Subject",
    "Status",
    "Message ID",
]


class ExcelTracker:
    def __init__(self, tracker_path: Path, columns: list[str] | None = None) -> None:
        self.tracker_path = tracker_path
        self.columns = columns or DEFAULT_COLUMNS

    def _ensure_workbook(self) -> None:
        self.tracker_path.parent.mkdir(parents=True, exist_ok=True)
        if self.tracker_path.exists():
            return
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Recruiter Tracker"
        sheet.append(self.columns)
        workbook.save(self.tracker_path)
        log.info("excel_tracker_created", path=str(self.tracker_path))

    def upsert_row(self, row: TrackerRow) -> None:
        self._ensure_workbook()
        workbook = load_workbook(self.tracker_path)
        sheet = workbook.active

        message_id_col = self.columns.index("Message ID") + 1
        existing_row_idx = None
        for idx in range(2, sheet.max_row + 1):
            if sheet.cell(row=idx, column=message_id_col).value == row.message_id:
                existing_row_idx = idx
                break

        values = [
            row.recruiter_name,
            row.contact_email,
            row.contact_phone or "",
            row.company,
            row.role_applied_for,
            row.match_score,
            _format_datetime(row.date_of_first_reply),
            row.email_subject,
            row.status.value,
            row.message_id,
        ]

        if existing_row_idx:
            for col_idx, value in enumerate(values, start=1):
                sheet.cell(row=existing_row_idx, column=col_idx, value=value)
            log.info("excel_tracker_updated", message_id=row.message_id)
        else:
            sheet.append(values)
            log.info("excel_tracker_appended", message_id=row.message_id)

        workbook.save(self.tracker_path)


def _format_datetime(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.strftime("%Y-%m-%d %H:%M:%S UTC")
